from __future__ import annotations

import io
import re
import traceback
from typing import Dict, List, Tuple

import pandas as pd
import streamlit as st


st.set_page_config(
    page_title="CSB Textdatei Tourzuordnung",
    page_icon="🚚",
    layout="wide",
)

st.title("🚚 CSB Textdatei Tourzuordnung")
st.caption(
    "Auswertung aus der CSB Textdatei. Schwerpunkt: Tour, Liefertag, Kundenname und CSB Nummer. "
    "Zeilen mit ????? werden ebenfalls der aktuellen Tour zugeordnet, CSB bleibt dann leer. Kundennamen wie TEST KUNDENNUMMER werden nicht mehr als Kopfzeile verworfen."
)

st.info(
    "CSB Textdatei hochladen. Die App liest je Tourblock die Kundenzeilen aus "
    "und erstellt eine Excel-Datei mit farblich gruppierten Tour-Endungen."
)


WOCHENTAG_MAP = {
    "montag": "Mo",
    "dienstag": "Die",
    "mittwoch": "Mitt",
    "donnerstag": "Don",
    "freitag": "Fr",
    "samstag": "Sam",
    "sonntag": "So",
}

TAG_REIHENFOLGE = ["Mo", "Die", "Mitt", "Don", "Fr", "Sam", "So"]

# Dezente Farben für gleiche Tour-Endungen, zum Beispiel 1999, 2999, 3999.
ENDUNG_FARBEN = [
    "FFF2CC",
    "D9EAD3",
    "D9EAF7",
    "EADCF8",
    "FCE4D6",
    "DDEBF7",
    "E2F0D9",
    "F4CCCC",
    "D9D2E9",
    "CFE2F3",
    "EADFCB",
    "E6F2E6",
]


def clean_text(value) -> str:
    if value is None:
        return ""
    value = str(value)
    value = value.replace("\x0c", " ")
    value = value.replace("\xa0", " ")
    value = value.replace("\x81", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip(" \t\r\n.;")


def norm_num(value) -> str:
    value = clean_text(value)
    if not value:
        return ""
    value = value.replace(",", ".")
    if re.fullmatch(r"\d+\.0", value):
        value = value[:-2]
    if re.fullmatch(r"\d+", value):
        return str(int(value))
    return value


def norm_tour(value) -> str:
    return norm_num(value)


def tour_endung(tour: str) -> str:
    tour = norm_tour(tour)
    if len(tour) >= 3 and tour.isdigit():
        return tour[-3:]
    return tour


def normalize_day(value: str) -> str:
    value_clean = clean_text(value).lower().replace(".", "")
    if value_clean in WOCHENTAG_MAP:
        return WOCHENTAG_MAP[value_clean]

    aliases = {
        "mo": "Mo",
        "mon": "Mo",
        "monday": "Mo",
        "die": "Die",
        "di": "Die",
        "dienst": "Die",
        "tuesday": "Die",
        "mitt": "Mitt",
        "mi": "Mitt",
        "mittw": "Mitt",
        "wednesday": "Mitt",
        "don": "Don",
        "do": "Don",
        "thursday": "Don",
        "fr": "Fr",
        "frei": "Fr",
        "friday": "Fr",
        "sam": "Sam",
        "sa": "Sam",
        "saturday": "Sam",
        "so": "So",
        "son": "So",
        "sunday": "So",
    }
    return aliases.get(value_clean, clean_text(value))


def decode_txt_bytes(data: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("cp1252", errors="replace")


def extract_customer_core(line: str):
    """
    Liest die wichtigste Information:
    - CSB Nummer, wenn vorhanden
    - Kundenname
    - Platzhalter-Zeilen mit ????? werden auch als Kunde übernommen,
      CSB Nummer bleibt dann leer.

    Bewusst nicht abhängig von Postleitzahl, Straße oder Punktspalten.
    """
    original = line.rstrip("\r\n").replace("\xa0", " ")

    if not original.strip():
        return None

    # Kopfzeilen, Summenzeilen und technische Druckzeilen nicht auswerten.
    # Wichtig: Nicht einfach nach "Kundennummer" irgendwo in der Zeile suchen,
    # weil es auch echte Kundennamen wie "TEST KUNDENNUMMER" geben kann.
    stripped = original.strip()

    header_patterns = [
        r"^Tour\s",
        r"^Tour-\s*/\s*Ladeplan",
        r"^Wochentag\s",
        r"^Fahrer\s*:",
        r"^LKW\s*:",
        r"^Tor\s*:",
        r"^km Stand",
        r"^Start Arbeitszeit",
        r"^Ende Arbeitszeit",
        r"^Druckdatum\s*:",
        r"^!!!!!",
        r"^La\.\s+Kunde",
        r"^Pa\s+",
        r"^von Zeit",
        r"^bis\s*$",
        r"^\d+\s+Anzahl Kunden\b",
    ]

    if any(re.search(pattern, stripped, re.IGNORECASE) for pattern in header_patterns):
        return None

    if "Rolli Rückgabe" in stripped or "Rolli Rueckgabe" in stripped:
        return None

    # Normale Kundenzeile:
    #   16968 Signature Foods ...
    #   1 13822 Kunde ...
    #
    # Sonderfall:
    #   ????? FABRIKVERKAUF FR. CE Am Heisterbusch ...
    #
    # Wichtig ist die Tourzuordnung. Bei ????? bleibt CSB leer.
    match = re.match(r"^\s*(?:(\d{1,3})\s+)?(?:(\d{3,6})|(\?{3,}))\s+(.*)$", original)
    if not match:
        return None

    ladereihenfolge = match.group(1) or ""
    csb_roh = match.group(2) or ""
    platzhalter = match.group(3) or ""
    rest = clean_text(match.group(4))

    csb = norm_num(csb_roh) if csb_roh else ""

    if not rest:
        return None

    # Punktspalten am Ende entfernen.
    rest = re.sub(r"(?:\s*\.){2,}\s*$", "", rest).strip()

    # Ab hier soll vor allem der Name sauber entstehen.
    # Adresse und Ort sind zweitrangig.
    name_basis = rest

    # Ortsangabe am Ende entfernen.
    # Unterstützt deutsche und ausländische Postleitzahlen wie A-4890 und NL-7580.
    location_match = re.search(
        r"\b(?:[A-Z]{1,3}-)?\d{4,5}\s+[A-ZÄÖÜa-zäöüß][A-ZÄÖÜa-zäöüß0-9 .\-/]*$",
        name_basis,
    )
    if location_match:
        name_basis = name_basis[:location_match.start()].rstrip()

    # Bekannte Straßenmuster entfernen, damit nur der Kundenname bleibt.
    street_match = re.search(
        r"\b("
        r"[A-ZÄÖÜa-zäöüß0-9.\-/]+(?:straße|strasse|str\.|str|weg|allee|damm|ring|platz|chaussee)"
        r"|Hauptstraße|Hauptstrasse|Hauptstr\.|Hauptstr"
        r"|Bahnhofstraße|Bahnhofstrasse|Bahnhofstr\.|Bahnhofstr"
        r"|Industriestraße|Industriestrasse|Industriestr\.|Industriestr"
        r"|Industrieterr\.?|INDUSTRIETERR\.?"
        r"|Am\s+[A-ZÄÖÜa-zäöüß]"
        r"|An\s+der\s+[A-ZÄÖÜa-zäöüß]"
        r"|Auf\s+dem\s+[A-ZÄÖÜa-zäöüß]"
        r")\b.*$",
        name_basis,
        flags=re.IGNORECASE,
    )
    if street_match and street_match.start() > 0:
        name = clean_text(name_basis[:street_match.start()])
    else:
        # Fester CSB Ausdruck: Name steht häufig am Anfang mit Abstand zur Straße.
        parts = [clean_text(part) for part in re.split(r"\s{2,}", name_basis) if clean_text(part)]
        if parts:
            name = parts[0]
        else:
            # Letzter Fallback: maximal die ersten 40 Zeichen als Name nehmen.
            name = clean_text(name_basis[:40])

    if not name:
        name = clean_text(rest[:40])

    return {
        "Ladereihenfolge aus Textdatei": ladereihenfolge,
        "CSB Nummer": csb,
        "Kunde": name,
        "CSB Platzhalter": "ja" if platzhalter else "",
        "Originalzeile": clean_text(original),
    }


def parse_csb_text(text: str) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    current_tour = ""
    current_day_raw = ""
    current_day = ""
    current_tour_text = ""
    position = 0

    rows: List[dict] = []
    tour_meta: Dict[str, dict] = {}

    tour_re = re.compile(r"^\s*Tour\s+(\d{3,6})\b(.*?)(?:LKW:|$)", re.IGNORECASE)
    day_re = re.compile(r"^\s*Wochentag\s+(.+?)(?:Fahrer:|$)", re.IGNORECASE)
    count_re = re.compile(r"^\s*(\d+)\s+Anzahl Kunden\b", re.IGNORECASE)

    for raw_line in text.splitlines():
        line = raw_line.rstrip("\r\n")

        day_match = day_re.search(line)
        if day_match:
            current_day_raw = clean_text(day_match.group(1))
            current_day = normalize_day(current_day_raw)

        tour_match = tour_re.search(line)
        if tour_match:
            current_tour = norm_tour(tour_match.group(1))
            current_tour_text = clean_text(tour_match.group(2))
            position = 0

            # Falls im Text kein Wochentag steht, wird der Liefertag aus der ersten Tourziffer abgeleitet.
            if current_tour and not current_day:
                current_day = {
                    "1": "Mo",
                    "2": "Die",
                    "3": "Mitt",
                    "4": "Don",
                    "5": "Fr",
                    "6": "Sam",
                    "7": "So",
                }.get(current_tour[0], "")

            tour_meta[current_tour] = {
                "Tour": current_tour,
                "Tour Endung": tour_endung(current_tour),
                "Liefertag": current_day,
                "Wochentag aus Textdatei": current_day_raw,
                "Tour Text": current_tour_text,
                "Soll Kunden laut Textdatei": None,
            }
            continue

        count_match = count_re.search(line)
        if count_match and current_tour:
            tour_meta.setdefault(
                current_tour,
                {
                    "Tour": current_tour,
                    "Tour Endung": tour_endung(current_tour),
                    "Liefertag": current_day,
                    "Wochentag aus Textdatei": current_day_raw,
                    "Tour Text": current_tour_text,
                    "Soll Kunden laut Textdatei": None,
                },
            )
            tour_meta[current_tour]["Soll Kunden laut Textdatei"] = int(count_match.group(1))
            continue

        customer = extract_customer_core(line)
        if customer and current_tour:
            position += 1
            rows.append(
                {
                    "Tour": current_tour,
                    "Tour Endung": tour_endung(current_tour),
                    "Liefertag": current_day,
                    "Position im Tourblock": position,
                    "Ladereihenfolge aus Textdatei": customer["Ladereihenfolge aus Textdatei"],
                    "CSB Nummer": customer["CSB Nummer"],
                    "Kunde": customer["Kunde"],
                    "CSB Platzhalter": customer["CSB Platzhalter"],
                    "Tour Text": current_tour_text,
                    "Wochentag aus Textdatei": current_day_raw,
                    "Originalzeile": customer["Originalzeile"],
                }
            )

    kunden_df = pd.DataFrame(rows)

    if kunden_df.empty:
        empty_touren = pd.DataFrame(
            columns=[
                "Tour",
                "Tour Endung",
                "Liefertag",
                "Wochentag aus Textdatei",
                "Tour Text",
                "Soll Kunden laut Textdatei",
                "Ausgelesene Kunden",
                "Differenz",
                "Status",
            ]
        )
        empty_doppelt = pd.DataFrame(
            columns=["Liefertag", "CSB Nummer", "Kunde", "Anzahl Touren", "Touren"]
        )
        return kunden_df, empty_touren, empty_doppelt

    erkannte = (
        kunden_df.groupby("Tour", as_index=False)
        .size()
        .rename(columns={"size": "Ausgelesene Kunden"})
    )

    touren_df = pd.DataFrame(tour_meta.values()).merge(erkannte, on="Tour", how="outer")
    touren_df["Soll Kunden laut Textdatei"] = pd.to_numeric(
        touren_df["Soll Kunden laut Textdatei"], errors="coerce"
    )
    touren_df["Ausgelesene Kunden"] = (
        pd.to_numeric(touren_df["Ausgelesene Kunden"], errors="coerce")
        .fillna(0)
        .astype(int)
    )
    touren_df["Differenz"] = touren_df["Ausgelesene Kunden"] - touren_df["Soll Kunden laut Textdatei"]

    def build_status(row) -> str:
        if pd.isna(row["Soll Kunden laut Textdatei"]):
            return "Keine Sollzahl gefunden"
        if row["Differenz"] == 0:
            return "OK"
        return "Abweichung"

    touren_df["Status"] = touren_df.apply(build_status, axis=1)

    # Doppelte nur für echte CSB Nummern prüfen. Platzhalter ohne CSB werden ignoriert.
    kunden_mit_csb = kunden_df[kunden_df["CSB Nummer"].astype(str).str.strip() != ""].copy()
    if kunden_mit_csb.empty:
        doppelt_df = pd.DataFrame(
            columns=["Liefertag", "CSB Nummer", "Kunde", "Anzahl Touren", "Touren"]
        )
    else:
        doppelt_df = (
            kunden_mit_csb.groupby(["Liefertag", "CSB Nummer"], as_index=False)
            .agg(
                Kunde=("Kunde", "first"),
                Anzahl_Touren=("Tour", "nunique"),
                Touren=("Tour", lambda values: ", ".join(sorted(set(map(str, values))))),
            )
        )
        doppelt_df = doppelt_df[doppelt_df["Anzahl_Touren"] > 1].copy()
        doppelt_df = doppelt_df.rename(columns={"Anzahl_Touren": "Anzahl Touren"})

    kunden_df = kunden_df.sort_values(
        ["Tour Endung", "Liefertag", "Tour", "Position im Tourblock"], kind="stable"
    ).reset_index(drop=True)

    touren_df = touren_df.sort_values(["Tour Endung", "Liefertag", "Tour"], kind="stable").reset_index(drop=True)

    if not doppelt_df.empty:
        doppelt_df = doppelt_df.sort_values(
            ["Liefertag", "CSB Nummer"], kind="stable"
        ).reset_index(drop=True)

    return kunden_df, touren_df, doppelt_df


def make_color_map(endungen: List[str]) -> Dict[str, str]:
    unique_endungen = sorted({clean_text(e) for e in endungen if clean_text(e)})
    return {
        endung: ENDUNG_FARBEN[index % len(ENDUNG_FARBEN)]
        for index, endung in enumerate(unique_endungen)
    }


def make_excel_export(
    kunden_df: pd.DataFrame,
    touren_df: pd.DataFrame,
    doppelt_df: pd.DataFrame,
) -> bytes:
    output = io.BytesIO()
    color_map = make_color_map(kunden_df["Tour Endung"].astype(str).tolist()) if not kunden_df.empty else {}

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        kunden_df.to_excel(writer, sheet_name="Tourzuordnung", index=False)
        touren_df.to_excel(writer, sheet_name="Touren Prüfung", index=False)
        doppelt_df.to_excel(writer, sheet_name="Doppelte Kunden je Tag", index=False)

        tages_df = pd.DataFrame()
        if not kunden_df.empty:
            tages_df = (
                kunden_df.groupby("Liefertag", as_index=False)
                .agg(
                    Anzahl_Lieferungen=("Kunde", "count"),
                    Anzahl_Kunden_mit_CSB=("CSB Nummer", lambda values: sum(1 for value in values if str(value).strip())),
                    Anzahl_Platzhalter=("CSB Platzhalter", lambda values: sum(1 for value in values if str(value).strip())),
                    Anzahl_Touren=("Tour", "nunique"),
                )
                .rename(
                    columns={
                        "Anzahl_Lieferungen": "Anzahl Lieferungen",
                        "Anzahl_Kunden_mit_CSB": "Anzahl Kunden mit CSB",
                        "Anzahl_Platzhalter": "Anzahl Platzhalter ohne CSB",
                        "Anzahl_Touren": "Anzahl Touren",
                    }
                )
            )
            tages_df["Sortierung"] = tages_df["Liefertag"].map(
                {tag: index for index, tag in enumerate(TAG_REIHENFOLGE)}
            )
            tages_df = (
                tages_df.sort_values("Sortierung", kind="stable")
                .drop(columns=["Sortierung"])
                .reset_index(drop=True)
            )

        tages_df.to_excel(writer, sheet_name="Tagesübersicht", index=False)

        from copy import copy
        from openpyxl.styles import Alignment, Font, PatternFill

        workbook = writer.book
        header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for worksheet in workbook.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

            for cell in worksheet[1]:
                cell.font = copy(header_font)
                cell.fill = copy(header_fill)
                cell.alignment = copy(header_alignment)

            # Farblich nach Tour-Endung markieren.
            if worksheet.title in ("Tourzuordnung", "Touren Prüfung"):
                header_names = [cell.value for cell in worksheet[1]]
                if "Tour Endung" in header_names:
                    endung_col = header_names.index("Tour Endung") + 1
                    for row in range(2, worksheet.max_row + 1):
                        endung = clean_text(worksheet.cell(row=row, column=endung_col).value)
                        color = color_map.get(endung)
                        if color:
                            fill = PatternFill(fill_type="solid", fgColor=color)
                            for col in range(1, worksheet.max_column + 1):
                                worksheet.cell(row=row, column=col).fill = fill

            for column in worksheet.columns:
                column_letter = column[0].column_letter
                max_length = 0
                for cell in column:
                    value = "" if cell.value is None else str(cell.value)
                    max_length = max(max_length, len(value))
                worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 70)

    return output.getvalue()


def style_by_tour_endung(df: pd.DataFrame):
    if df.empty or "Tour Endung" not in df.columns:
        return df

    color_map = make_color_map(df["Tour Endung"].astype(str).tolist())

    def apply_row(row):
        endung = clean_text(row.get("Tour Endung", ""))
        color = color_map.get(endung)
        if not color:
            return [""] * len(row)
        return [f"background-color: #{color}" for _ in row]

    return df.style.apply(apply_row, axis=1)


def show_dataframe(df: pd.DataFrame, farbig: bool = False):
    if farbig:
        st.dataframe(style_by_tour_endung(df), use_container_width=True, hide_index=True)
    else:
        st.dataframe(df, use_container_width=True, hide_index=True)


uploaded_txt = st.file_uploader("CSB Textdatei hochladen", type=["txt"])

if uploaded_txt is None:
    st.stop()

try:
    with st.spinner("Textdatei wird ausgewertet..."):
        txt_text = decode_txt_bytes(uploaded_txt.getvalue())
        kunden_df, touren_df, doppelt_df = parse_csb_text(txt_text)

    if kunden_df.empty:
        st.error("Es wurden keine Kundenzeilen oder Platzhalter-Zeilen erkannt.")
        st.stop()

    auffaellige_touren = int((touren_df["Status"] != "OK").sum()) if not touren_df.empty else 0
    platzhalter_anzahl = int((kunden_df["CSB Platzhalter"].astype(str).str.strip() != "").sum())
    kunden_mit_csb = int((kunden_df["CSB Nummer"].astype(str).str.strip() != "").sum())

    col1, col2, col3, col4, col5 = st.columns(5)
    col1.metric("Lieferungen", f"{len(kunden_df):,}".replace(",", "."))
    col2.metric("mit CSB", f"{kunden_mit_csb:,}".replace(",", "."))
    col3.metric("ohne CSB", f"{platzhalter_anzahl:,}".replace(",", "."))
    col4.metric("Touren", f"{kunden_df['Tour'].nunique():,}".replace(",", "."))
    col5.metric("Auffällige Touren", f"{auffaellige_touren:,}".replace(",", "."))

    if platzhalter_anzahl:
        st.info(f"{platzhalter_anzahl} Zeilen mit ????? wurden der jeweiligen Tour zugeordnet. Die CSB Nummer bleibt leer.")

    if auffaellige_touren == 0:
        st.success("Die ausgelesene Kundenzahl passt bei allen Touren zur Sollzahl aus der Textdatei.")
    else:
        st.warning("Es gibt Touren mit abweichender Kundenzahl oder ohne gefundene Sollzahl.")

    excel_bytes = make_excel_export(kunden_df, touren_df, doppelt_df)

    st.download_button(
        "Excel Export herunterladen",
        data=excel_bytes,
        file_name="CSB_Tourzuordnung_aus_Textdatei.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    tab1, tab2, tab3, tab4 = st.tabs(
        [
            "Tourzuordnung",
            "Touren Prüfung",
            "Auffällige Touren",
            "Doppelte Kunden je Tag",
        ]
    )

    with tab1:
        st.caption(
            "Wichtigste Auswertung: Tour, Tour-Endung, Liefertag, CSB Nummer und Kundenname. "
            "Gleiche Tour-Endungen sind farblich gleich markiert."
        )
        view_df = kunden_df[
            [
                "Tour",
                "Tour Endung",
                "Liefertag",
                "Position im Tourblock",
                "CSB Nummer",
                "Kunde",
                "CSB Platzhalter",
                "Originalzeile",
            ]
        ].copy()
        show_dataframe(view_df, farbig=True)

    with tab2:
        show_dataframe(touren_df, farbig=True)

    with tab3:
        auffaellig_df = touren_df[touren_df["Status"] != "OK"].copy()
        if auffaellig_df.empty:
            st.success("Keine auffälligen Touren gefunden.")
        else:
            show_dataframe(auffaellig_df, farbig=True)

    with tab4:
        if doppelt_df.empty:
            st.success("Keine doppelten Kunden je Liefertag gefunden.")
        else:
            show_dataframe(doppelt_df)

except Exception:
    st.error("Fehler beim Verarbeiten der Textdatei.")
    st.code(traceback.format_exc())
